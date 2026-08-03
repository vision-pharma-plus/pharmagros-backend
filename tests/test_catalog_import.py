"""
Bulk catalogue import from a spreadsheet.

The cases here are the ones that decide whether an onboarding import is
trustworthy: that a French-formatted number is read as a number, that a bad
row is reported by its Excel row number instead of aborting the file, and
above all that a file containing any error writes nothing at all. A partial
import is the failure mode that costs a pharmacy a day of reconciliation.
"""

from __future__ import annotations

import io
from decimal import Decimal

import pytest
from openpyxl import Workbook, load_workbook

from apps.catalog import imports
from apps.catalog.models import Manufacturer, Medicine


@pytest.fixture
def manufacturer(db):
    return Manufacturer.objects.create(code="SANOFI", name="Sanofi")


@pytest.fixture
def refs(category, unit, manufacturer):
    """The reference records an imported row points at by code."""
    return category, unit, manufacturer


def _workbook(rows, headers=None, name="import.xlsx") -> io.BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers or [column.header_fr for column in imports.COLUMNS])
    for row in rows:
        sheet.append(row)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    # The parser branches on the file extension, which on a real upload comes
    # from the multipart filename.
    buffer.name = name
    return buffer


def _row(**overrides) -> list:
    """A minimally valid row, as a list in template column order."""
    values = {
        "name": "Paracétamol",
        "category": "TEST",
        "unit_of_measure": "BTE",
        "unit_cost": "1200",
        "selling_price": "1800",
    }
    values.update(overrides)
    return [values.get(column.key, "") for column in imports.COLUMNS]


class TestTemplate:
    def test_lists_the_codes_that_currently_exist(self, refs):
        """
        The reference tab is the point of generating the template server-side.

        A static template can describe the *shape* of a category code but not
        which ones this database will accept, which is the part users get wrong.
        """
        workbook = load_workbook(io.BytesIO(imports.build_template(language="fr")))
        reference = workbook[workbook.sheetnames[1]]
        codes = [cell.value for cell in reference["A"]]

        assert "TEST" in codes
        assert "BTE" in codes
        assert "SANOFI" in codes

    def test_headers_match_what_the_parser_accepts(self, refs):
        """The template must round-trip: whatever it emits, the parser reads."""
        workbook = load_workbook(io.BytesIO(imports.build_template(language="fr")))
        headers = [cell.value for cell in workbook["Import"][1]]

        report = imports.parse_workbook(_workbook([_row()], headers=headers))

        assert report.fatal_error == ""
        assert not report.error_rows


class TestParsing:
    def test_valid_row_is_created(self, refs, admin_user):
        report = imports.parse_workbook(_workbook([_row()]))
        assert not report.error_rows

        imports.commit(report, actor=admin_user)

        assert report.created == 1
        medicine = Medicine.objects.get(name_fr="Paracétamol")
        assert medicine.selling_price == Decimal("1800")
        # Allocated by the numbering service, exactly as on the product form.
        assert medicine.product_code
        # And the opening price-history row the form would also have written.
        assert medicine.price_history.count() == 1

    def test_french_number_formats(self, refs):
        """
        Excel on a French locale writes "1 200,50", and pasted prices carry
        their unit. Rejecting either would fail on the majority of real files.
        """
        report = imports.parse_workbook(
            _workbook([_row(unit_cost="1 200,50", vat_rate="18%")])
        )

        assert not report.error_rows
        assert report.rows[0].data["unit_cost"] == Decimal("1200.50")
        assert report.rows[0].data["vat_rate"] == Decimal("18")

    def test_english_headers_are_accepted(self, refs):
        report = imports.parse_workbook(
            _workbook(
                [_row()], headers=[column.header_en for column in imports.COLUMNS]
            )
        )
        assert not report.error_rows

    def test_unknown_columns_are_ignored(self, refs):
        """A legacy export carries columns this catalogue has no field for."""
        headers = [column.header_fr for column in imports.COLUMNS] + ["Ancien code SAP"]
        report = imports.parse_workbook(_workbook([_row() + ["X-99"]], headers=headers))
        assert not report.error_rows

    def test_blank_rows_are_skipped(self, refs):
        """Trailing empty rows are normal in a hand-edited sheet."""
        report = imports.parse_workbook(
            _workbook([_row(), [""] * len(imports.COLUMNS)])
        )
        assert len(report.rows) == 1

    def test_csv_with_semicolons(self, refs):
        headers = ";".join(column.header_fr for column in imports.COLUMNS)
        line = ";".join(str(value) for value in _row())
        buffer = io.BytesIO(f"{headers}\n{line}\n".encode("utf-8-sig"))
        buffer.name = "catalogue.csv"

        report = imports.parse_workbook(buffer)

        assert not report.error_rows
        assert report.rows[0].data["name"] == "Paracétamol"

    def test_yes_no_columns(self, refs):
        report = imports.parse_workbook(
            _workbook([_row(is_controlled="OUI", requires_prescription="yes")])
        )

        assert not report.error_rows
        assert report.rows[0].data["is_controlled"] is True
        assert report.rows[0].data["requires_prescription"] is True


class TestRowErrors:
    def test_unknown_category_reports_the_excel_row_number(self, refs):
        report = imports.parse_workbook(_workbook([_row(), _row(category="NOPE")]))

        assert len(report.valid_rows) == 1
        error = report.error_rows[0]
        assert "category" in error.errors
        # Row 3: the header is row 1 and the first data row is row 2, so the
        # number points at the row the user can see and fix.
        assert error.row_number == 3

    def test_invalid_choice_lists_the_valid_codes(self, refs):
        report = imports.parse_workbook(_workbook([_row(dosage_form="PILL")]))
        assert "TABLET" in report.error_rows[0].errors["dosage_form"][0]

    def test_non_numeric_price(self, refs):
        report = imports.parse_workbook(_workbook([_row(selling_price="à voir")]))
        assert "selling_price" in report.error_rows[0].errors

    def test_max_stock_below_reorder_level(self, refs):
        """
        Mirrors the model check constraint, which would otherwise surface as an
        opaque IntegrityError partway through the write.
        """
        report = imports.parse_workbook(
            _workbook([_row(reorder_level="100", max_stock_level="50")])
        )
        assert "max_stock_level" in report.error_rows[0].errors

    def test_duplicate_product_code_within_the_file(self, refs):
        report = imports.parse_workbook(
            _workbook([_row(product_code="P1"), _row(product_code="P1")])
        )
        assert "product_code" in report.error_rows[0].errors

    def test_barcode_already_used_by_another_product(self, refs, admin_user):
        imports.commit(
            imports.parse_workbook(_workbook([_row(barcode="123456")])),
            actor=admin_user,
        )
        report = imports.parse_workbook(_workbook([_row(barcode="123456")]))
        assert "barcode" in report.error_rows[0].errors


class TestFatalErrors:
    def test_missing_required_column(self, refs):
        headers = [
            column.header_fr for column in imports.COLUMNS if column.key != "category"
        ]
        workbook = Workbook()
        workbook.active.append(headers)
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        buffer.name = "x.xlsx"

        report = imports.parse_workbook(buffer)

        assert "Code catégorie" in report.fatal_error

    def test_unreadable_file(self, refs):
        buffer = io.BytesIO(b"this is not a spreadsheet")
        buffer.name = "broken.xlsx"

        report = imports.parse_workbook(buffer)

        assert report.fatal_error
        assert not report.rows


class TestReimport:
    def test_existing_product_code_updates_in_place(self, refs, admin_user):
        """
        Re-importing a corrected file is how a catalogue gets fixed, so a
        known product code is an update, not a duplicate-key error.
        """
        imports.commit(
            imports.parse_workbook(_workbook([_row(product_code="P1")])),
            actor=admin_user,
        )

        report = imports.parse_workbook(
            _workbook([_row(product_code="P1", name="Doliprane", selling_price="2500")])
        )
        assert report.rows[0].existing_id

        imports.commit(report, actor=admin_user)

        assert report.updated == 1
        assert Medicine.objects.count() == 1
        medicine = Medicine.objects.get()
        assert medicine.name_fr == "Doliprane"
        assert medicine.selling_price == Decimal("2500")
        # The repricing is recorded, so an imported price change is as
        # auditable as one made on the screen.
        assert medicine.price_history.count() == 2


class TestImportApi:
    ENDPOINT = "/api/v1/catalog/medicines/import/"

    def _upload(self, client, rows, **extra):
        from django.core.files.uploadedfile import SimpleUploadedFile

        buffer = _workbook(rows)
        upload = SimpleUploadedFile(
            "import.xlsx",
            buffer.getvalue(),
            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
        )
        return client.post(
            self.ENDPOINT, {"file": upload, **extra}, format="multipart"
        )

    def test_template_downloads_as_a_workbook(self, auth_client, pharmacist):
        response = auth_client(pharmacist).get(
            "/api/v1/catalog/medicines/import-template/"
        )

        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]
        assert "attachment" in response["Content-Disposition"]

    def test_dry_run_writes_nothing(self, auth_client, admin_user, refs):
        response = self._upload(auth_client(admin_user), [_row()])

        assert response.status_code == 200
        assert response.data["valid_count"] == 1
        assert response.data["committed"] is False
        assert Medicine.objects.count() == 0

    def test_commit_creates_the_products(self, auth_client, admin_user, refs):
        response = self._upload(
            auth_client(admin_user), [_row(), _row(name="Ibuprofène")], dry_run="false"
        )

        assert response.status_code == 200
        assert response.data["committed"] is True
        assert response.data["created"] == 2
        assert Medicine.objects.count() == 2

    def test_a_file_with_any_error_writes_nothing(self, auth_client, admin_user, refs):
        """
        The whole point of the atomic commit: an operator must never be left
        reconciling which half of their spreadsheet landed.
        """
        response = self._upload(
            auth_client(admin_user),
            [_row(), _row(category="NOPE")],
            dry_run="false",
        )

        assert response.status_code == 200
        assert response.data["committed"] is False
        assert response.data["error_count"] == 1
        assert response.data["errors"][0]["row_number"] == 3
        assert Medicine.objects.count() == 0

    def test_wrong_file_type_is_rejected(self, auth_client, admin_user):
        from django.core.files.uploadedfile import SimpleUploadedFile

        upload = SimpleUploadedFile(
            "catalogue.pdf", b"%PDF-1.4", content_type="application/pdf"
        )
        response = auth_client(admin_user).post(
            self.ENDPOINT, {"file": upload}, format="multipart"
        )

        assert response.status_code == 400

    def test_requires_the_create_permission(self, auth_client, auditor, refs):
        """An auditor may read the catalogue but not bulk-create in it."""
        response = self._upload(auth_client(auditor), [_row()])
        assert response.status_code == 403
