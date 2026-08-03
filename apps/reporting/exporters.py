"""
Report export to CSV, Excel and PDF.

Two details that matter for this deployment:

  * CSV is written with a UTF-8 BOM and semicolon delimiter. Excel on a
    French/Belgian locale — the norm in Burundi — treats the comma as a
    decimal separator and will not split comma-delimited columns. Without the
    BOM it also mangles accented characters. Both defaults are wrong for the
    users who will actually open these files.

  * Decimals are written as numbers, not strings, so totals can be summed in
    the spreadsheet without retyping.

PDF is a third, non-negotiable format: it is what gets printed, signed and
filed. It shares the invoice PDF engine (WeasyPrint) and carries the company
logo, so an exported report is recognisably a company document.
"""

from __future__ import annotations

import csv
import io
from decimal import Decimal

from django.http import HttpResponse
from django.utils import timezone, translation

from apps.core.money import format_number

# Columns whose figures are not whole francs. A unit cost or a tax rate printed
# to zero decimals is what makes a report fail its own arithmetic: the reader
# multiplies the rounded figure they can see and gets a different answer from
# the value beside it. Quantities are dispensing amounts, which are routinely
# fractional.
#
# `unit_cost` is deliberately absent. The valuation and expiry reports already
# round it to whole francs before it reaches here — that rounding is what makes
# their lines multiply out — so forcing decimals would print ",0000" on every
# row and imply a precision the column no longer carries. The movement ledger
# is the one report whose unit cost is the raw four-decimal figure, and its
# rows are individually valued rather than re-derived by the reader.
_DECIMAL_PLACES = {
    "unit_price": 4,
    "quantity": 3,
    "quantity_delta": 3,
    "balance_after": 3,
    "tax_rate": 2,
    "discount_percent": 2,
    "margin_percent": 2,
}


def _export_locale() -> str:
    """
    The language this export is being rendered in.

    Read from Django's active translation, which the request middleware has
    already set from the `X-Language` header — so the numbers in a file follow
    the same language as its translated column headers.
    """
    return "en" if (translation.get_language() or "fr").startswith("en") else "fr"


def _stringify(value, *, locale: str = "fr", decimals: int = 0):
    if isinstance(value, Decimal):
        return format_number(value, locale=locale, decimals=decimals)
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    if value is None:
        return ""
    return str(value)


def to_csv(rows: list[dict], columns: list[tuple[str, str]], filename: str) -> HttpResponse:
    """
    Render rows to a CSV download.

    `columns` is a list of (key, header) pairs, so the export controls both
    column order and the translated header text.
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    locale = _export_locale()

    writer.writerow([header for _key, header in columns])
    for row in rows:
        writer.writerow(
            [
                _stringify(
                    row.get(key), locale=locale, decimals=_DECIMAL_PLACES.get(key, 0),
                )
                for key, _header in columns
            ]
        )

    # utf-8-sig emits the BOM Excel needs to detect UTF-8.
    response = HttpResponse(
        buffer.getvalue().encode("utf-8-sig"), content_type="text/csv; charset=utf-8"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    return response


def to_excel(
    rows: list[dict],
    columns: list[tuple[str, str]],
    filename: str,
    *,
    title: str = "",
    metadata: dict | None = None,
) -> HttpResponse:
    """Render rows to an .xlsx download with a header block and frozen panes."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    locale = _export_locale()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (title or "Rapport")[:31]  # Excel caps sheet names at 31 chars

    row_index = 1
    if title:
        sheet.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        row_index = 2

    if metadata:
        for key, value in metadata.items():
            sheet.cell(row=row_index, column=1, value=f"{key}:").font = Font(bold=True)
            sheet.cell(row=row_index, column=2, value=_stringify(value, locale=locale))
            row_index += 1

    sheet.cell(row=row_index, column=1, value="Généré le / Generated:").font = Font(bold=True)
    sheet.cell(row=row_index, column=2, value=timezone.localtime().strftime("%d/%m/%Y %H:%M"))
    row_index += 2

    header_row = row_index
    header_fill = PatternFill("solid", fgColor="0F5132")
    header_font = Font(bold=True, color="FFFFFF")

    for column_index, (_key, header) in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for offset, row in enumerate(rows, start=1):
        for column_index, (key, _header) in enumerate(columns, start=1):
            value = row.get(key)
            if isinstance(value, Decimal):
                # Written as a real number so the spreadsheet can total it.
                # Excel applies the reader's own locale separators to a numeric
                # cell, so only the decimal count is set here — forcing "#,##0"
                # on a unit cost hid the decimals that make the row add up.
                cell = sheet.cell(row=header_row + offset, column=column_index, value=float(value))
                decimals = _DECIMAL_PLACES.get(key, 0)
                cell.number_format = "#,##0" if decimals == 0 else f"#,##0.{'0' * decimals}"
            elif hasattr(value, "strftime"):
                cell = sheet.cell(row=header_row + offset, column=column_index, value=value)
                cell.number_format = "DD/MM/YYYY"
            else:
                sheet.cell(
                    row=header_row + offset, column=column_index,
                    value=_stringify(value, locale=locale),
                )

    # Width from content, capped so a long note does not create a 200-char column.
    for column_index, (key, header) in enumerate(columns, start=1):
        longest = max(
            [len(header)]
            + [
                len(_stringify(row.get(key), locale=locale, decimals=_DECIMAL_PLACES.get(key, 0)))
                for row in rows[:200]
            ]
            or [10]
        )
        sheet.column_dimensions[get_column_letter(column_index)].width = min(longest + 3, 45)

    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


def _pdf_cell(value, *, locale: str = "fr", decimals: int = 0) -> dict:
    """
    Render one cell for the PDF grid.

    Numbers are grouped and right-aligned, which is how a BIF figure is read
    here; unlike the spreadsheet export there is nothing to re-sum downstream,
    so display beats machine-readability.

    `decimals` is what keeps the printed page self-consistent: a unit cost
    forced to whole francs no longer multiplies out against the value beside
    it, which is exactly the discrepancy reported on the valuation report.
    """
    if isinstance(value, Decimal):
        return {
            "value": format_number(value, locale=locale, decimals=decimals),
            "numeric": True,
        }
    if isinstance(value, int) and not isinstance(value, bool):
        return {"value": format_number(value, locale=locale), "numeric": True}
    return {"value": _stringify(value, locale=locale), "numeric": False}


def to_pdf(
    rows: list[dict],
    columns: list[tuple[str, str]],
    filename: str,
    *,
    title: str = "",
    metadata: dict | None = None,
) -> HttpResponse:
    """
    Render rows to a printable PDF download carrying the company header.

    The template's `@page` footer and column headers are translated, so the
    active language at render time decides the document's language — the same
    contract the invoice renderer follows.
    """
    from django.conf import settings
    from django.template.loader import render_to_string
    from django.utils.translation import get_language

    from apps.invoicing.pdf import _html_to_pdf

    locale = _export_locale()

    html = render_to_string(
        "pdf/report.html",
        {
            "title": title or filename,
            "company": settings.COMPANY,
            "generated_at": timezone.localtime().strftime("%d/%m/%Y %H:%M"),
            # The template sets <html lang> and the footer from this; without
            # it the attribute renders empty and WeasyPrint has no language to
            # hand Pango for hyphenation.
            "LANGUAGE_CODE": get_language() or "fr",
            "headers": [header for _key, header in columns],
            # Pre-flattened to a list of cells per row: the template engine
            # cannot index a dict by a loop variable.
            "rows": [
                [
                    _pdf_cell(
                        row.get(key),
                        locale=locale,
                        decimals=_DECIMAL_PLACES.get(key, 0),
                    )
                    for key, _header in columns
                ]
                for row in rows
            ],
            "row_count": len(rows),
            # Pairs, not a dict, so the template can unpack label/value.
            # Formatted through _pdf_cell so a metadata total is grouped the
            # same way as the column it summarises.
            "metadata": [
                (key, _pdf_cell(value, locale=locale)["value"])
                for key, value in (metadata or {}).items()
            ],
        },
    )

    response = HttpResponse(_html_to_pdf(html), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response


def export(
    fmt: str,
    rows: list[dict],
    columns: list[tuple[str, str]],
    filename: str,
    *,
    title: str = "",
    metadata: dict | None = None,
) -> HttpResponse:
    """Dispatch to the requested format."""
    if fmt == "xlsx":
        return to_excel(rows, columns, filename, title=title, metadata=metadata)
    if fmt == "pdf":
        return to_pdf(rows, columns, filename, title=title, metadata=metadata)
    return to_csv(rows, columns, filename)
