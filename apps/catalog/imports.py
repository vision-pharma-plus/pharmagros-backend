"""
Bulk catalogue import from a spreadsheet.

An existing pharmacy arrives with its product list in Excel, not in this
database. Retyping several thousand rows through the product form is not a
realistic onboarding path, so the catalogue accepts a workbook directly.

Three decisions shape this module:

  * **The template is generated, not documented.** A prose description of the
    expected columns is something a user has to reproduce by hand, and every
    hand-built file differs. `build_template()` emits the exact sheet the
    parser reads, with the reference codes for categories, units and
    manufacturers embedded on a second tab — so the file the user fills in is
    the file the importer expects.

  * **Validation is a separate pass from writing.** `parse_workbook()` returns
    every row's outcome without touching the database, which is what lets the
    UI show "31 rows ready, 4 in error" before anything is committed. A partial
    import that stops at row 200 leaves the operator with no idea which half
    of their catalogue landed.

  * **Rows are validated through the ordinary serializer and written through
    the ordinary service.** Import is a bulk front door onto product creation,
    not a second implementation of it — the price-history row, the product-code
    allocation and the audit entry must be identical whether a product arrives
    through the form or through a spreadsheet.

CSV is accepted alongside .xlsx because exporting to CSV is often the only
thing a legacy system can do. It is read with the same semicolon/UTF-8-BOM
convention the report exporter writes, since that is what Excel on a French
locale produces.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction
from django.utils.translation import gettext as _

from .models import (
    Category,
    DosageForm,
    Manufacturer,
    Medicine,
    ProductStatus,
    StorageCondition,
    UnitOfMeasure,
)

# Excel's own ceiling is 1,048,576 rows; the practical limit here is the
# request timeout, and a catalogue larger than this is a data-migration job
# rather than a screen action.
MAX_ROWS = 5000


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ImportColumn:
    """One spreadsheet column and how it maps onto the serializer."""

    key: str
    header_fr: str
    header_en: str
    required: bool = False
    hint_fr: str = ""
    hint_en: str = ""
    example: str = ""
    # Width in Excel character units — set per column so the generated
    # template is readable without the user resizing anything.
    width: int = 18

    def header(self, language: str) -> str:
        return self.header_en if language.startswith("en") else self.header_fr

    def hint(self, language: str) -> str:
        return self.hint_en if language.startswith("en") else self.hint_fr


# Order here is the column order in the template. Identity and classification
# come first because they are what the operator is copying from their old
# system; the optional stock-policy and regulatory columns trail behind so a
# minimal file can simply stop after the prices.
COLUMNS: list[ImportColumn] = [
    ImportColumn(
        "product_code", "Code produit", "Product code",
        hint_fr="Laissez vide pour une attribution automatique.",
        hint_en="Leave blank to allocate automatically.",
        example="MED-000123", width=16,
    ),
    ImportColumn(
        "name", "Dénomination", "Name", required=True,
        hint_fr="Nom commercial du produit.",
        hint_en="Commercial name of the product.",
        example="Paracétamol", width=28,
    ),
    ImportColumn(
        "generic_name", "DCI", "Generic name (INN)",
        hint_fr="Dénomination commune internationale.",
        hint_en="International non-proprietary name.",
        example="Paracetamol", width=24,
    ),
    ImportColumn(
        "brand_name", "Marque", "Brand name",
        example="Doliprane", width=20,
    ),
    ImportColumn(
        "strength", "Dosage", "Strength",
        hint_fr="Par exemple : 500 mg, 250 mg/5 ml.",
        hint_en="For example: 500 mg, 250 mg/5 ml.",
        example="500 mg", width=14,
    ),
    ImportColumn(
        "dosage_form", "Forme galénique", "Dosage form",
        hint_fr="Code de la liste « Références ». Par défaut : TABLET.",
        hint_en="Code from the 'Reference' tab. Defaults to TABLET.",
        example="TABLET", width=18,
    ),
    ImportColumn(
        "pack_size", "Conditionnement", "Pack size",
        example="Boîte de 100", width=18,
    ),
    ImportColumn(
        "category", "Code catégorie", "Category code", required=True,
        hint_fr="Code d'une catégorie existante (onglet « Références »).",
        hint_en="Code of an existing category (see the 'Reference' tab).",
        example="ANALG", width=16,
    ),
    ImportColumn(
        "manufacturer", "Code fabricant", "Manufacturer code",
        hint_fr="Code d'un fabricant existant. Facultatif.",
        hint_en="Code of an existing manufacturer. Optional.",
        example="SANOFI", width=16,
    ),
    ImportColumn(
        "unit_of_measure", "Code unité", "Unit code", required=True,
        hint_fr="Code d'une unité existante (onglet « Références »).",
        hint_en="Code of an existing unit (see the 'Reference' tab).",
        example="BTE", width=14,
    ),
    ImportColumn(
        "unit_cost", "Coût de référence HT", "Reference unit cost (excl. VAT)",
        required=True,
        hint_fr="Prix d'achat hors taxe. Chiffres uniquement.",
        hint_en="Purchase price excluding tax. Digits only.",
        example="1200", width=20,
    ),
    ImportColumn(
        "selling_price", "Prix de vente TTC", "Selling price (VAT incl.)",
        required=True,
        hint_fr="Prix payé par le client, TVA comprise.",
        hint_en="Price the customer pays, VAT included.",
        example="1800", width=20,
    ),
    ImportColumn(
        "wholesale_price", "Prix de gros TTC", "Wholesale price (VAT incl.)",
        hint_fr="Prix institutionnel, TVA comprise. Facultatif.",
        hint_en="Institutional price, VAT included. Optional.",
        example="1500", width=20,
    ),
    ImportColumn(
        "vat_rate", "Taux de TVA (%)", "VAT rate (%)",
        hint_fr="Par défaut : 18. Mettez 0 pour un produit exonéré.",
        hint_en="Defaults to 18. Use 0 for an exempt product.",
        example="18", width=16,
    ),
    ImportColumn(
        "is_vat_exempt", "Exonéré de TVA", "VAT exempt",
        hint_fr="OUI ou NON. Par défaut : NON.",
        hint_en="YES or NO. Defaults to NO.",
        example="NON", width=16,
    ),
    ImportColumn(
        "reorder_level", "Seuil de réappro.", "Reorder level",
        example="50", width=18,
    ),
    ImportColumn(
        "safety_stock", "Stock de sécurité", "Safety stock",
        example="20", width=18,
    ),
    ImportColumn(
        "max_stock_level", "Stock maximum", "Maximum stock level",
        hint_fr="Doit être supérieur au seuil de réapprovisionnement, ou 0.",
        hint_en="Must be above the reorder level, or 0.",
        example="500", width=18,
    ),
    ImportColumn(
        "expiry_alert_days", "Alerte péremption (j)", "Expiry alert (days)",
        hint_fr="Par défaut : 180.",
        hint_en="Defaults to 180.",
        example="180", width=20,
    ),
    ImportColumn(
        "storage_condition", "Conservation", "Storage condition",
        hint_fr="Code de la liste « Références ». Par défaut : AMBIENT.",
        hint_en="Code from the 'Reference' tab. Defaults to AMBIENT.",
        example="AMBIENT", width=18,
    ),
    ImportColumn(
        "requires_prescription", "Sur ordonnance", "Prescription required",
        hint_fr="OUI ou NON. Par défaut : NON.",
        hint_en="YES or NO. Defaults to NO.",
        example="NON", width=18,
    ),
    ImportColumn(
        "is_controlled", "Substance contrôlée", "Controlled substance",
        hint_fr="OUI ou NON. Par défaut : NON.",
        hint_en="YES or NO. Defaults to NO.",
        example="NON", width=20,
    ),
    ImportColumn(
        "atc_code", "Code ATC", "ATC code",
        example="N02BE01", width=14,
    ),
    ImportColumn(
        "registration_number", "Numéro d'AMM", "Marketing authorisation no.",
        example="AMM-2024-001", width=22,
    ),
    ImportColumn(
        "barcode", "Code-barres", "Barcode",
        hint_fr="Doit être unique s'il est renseigné.",
        hint_en="Must be unique when supplied.",
        example="5901234123457", width=18,
    ),
    ImportColumn(
        "notes", "Notes", "Notes",
        example="", width=30,
    ),
]

COLUMNS_BY_KEY = {column.key: column for column in COLUMNS}

# Numeric columns that must parse as a decimal, with the default applied when
# the cell is blank. Kept separate from the column list because the parsing
# rule, not the presentation, is what varies between them.
DECIMAL_DEFAULTS: dict[str, Decimal] = {
    "unit_cost": Decimal("0"),
    "selling_price": Decimal("0"),
    "wholesale_price": Decimal("0"),
    "vat_rate": Decimal("18"),
    "reorder_level": Decimal("0"),
    "safety_stock": Decimal("0"),
    "max_stock_level": Decimal("0"),
}

BOOLEAN_COLUMNS = ("is_vat_exempt", "requires_prescription", "is_controlled")

# Accepted spellings for a yes/no cell. Both languages are listed because the
# template ships localised examples, and 1/0 because that is what a CSV
# exported from an older system usually contains.
TRUE_TOKENS = {"oui", "yes", "y", "o", "true", "vrai", "1", "x"}
FALSE_TOKENS = {"non", "no", "n", "false", "faux", "0", ""}


# ---------------------------------------------------------------------------
# Result types
# ---------------------------------------------------------------------------


@dataclass
class RowResult:
    """Outcome for a single spreadsheet row."""

    # 1-based row number as displayed by Excel, so an error message points at
    # the row the user can actually see and fix.
    row_number: int
    data: dict = field(default_factory=dict)
    errors: dict[str, list[str]] = field(default_factory=dict)
    # Set when the row matches an existing product; drives update-vs-create.
    existing_id: str | None = None

    @property
    def is_valid(self) -> bool:
        return not self.errors


@dataclass
class ImportReport:
    """Aggregate outcome, shaped for direct serialisation to the client."""

    rows: list[RowResult] = field(default_factory=list)
    created: int = 0
    updated: int = 0
    # File-level failures (unreadable workbook, missing columns) that make the
    # row list meaningless.
    fatal_error: str = ""

    @property
    def valid_rows(self) -> list[RowResult]:
        return [row for row in self.rows if row.is_valid]

    @property
    def error_rows(self) -> list[RowResult]:
        return [row for row in self.rows if not row.is_valid]

    def as_dict(self, *, limit_errors: int = 200) -> dict:
        return {
            "total_rows": len(self.rows),
            "valid_count": len(self.valid_rows),
            "error_count": len(self.error_rows),
            "created": self.created,
            "updated": self.updated,
            "fatal_error": self.fatal_error,
            "errors": [
                {
                    "row_number": row.row_number,
                    # Echoed so the client can label the row with something
                    # recognisable rather than only a line number.
                    "name": row.data.get("name", "") or row.data.get("product_code", ""),
                    "errors": row.errors,
                }
                for row in self.error_rows[:limit_errors]
            ],
        }


# ---------------------------------------------------------------------------
# Cell coercion
# ---------------------------------------------------------------------------


def _clean(value) -> str:
    """Normalise a raw cell to a trimmed string."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        # openpyxl reads every unformatted number as a float, which would turn
        # the pack size 100 into "100.0" and a barcode into scientific
        # notation. Integers are restored to their integer spelling.
        return str(int(value))
    return str(value).strip()


def _to_decimal(raw: str) -> Decimal:
    """
    Parse a money or quantity cell.

    Handles the separators the local Excel produces: a comma decimal mark, and
    thin/non-breaking spaces used as thousands separators. A value typed as
    "1 200,50" is the same number as "1200.50" and must not be an error.
    """
    text = (
        raw.replace(" ", "")
        .replace(" ", "")
        .replace(" ", "")
        .replace("'", "")
    )
    # A comma is a decimal mark here, never a thousands separator — but a file
    # from an anglophone system may use commas for thousands, in which case
    # there is more than one and none of them is decimal.
    if text.count(",") > 1:
        text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    # Currency symbols and codes travel with pasted values.
    for token in ("BIF", "FBu", "Fbu", "€", "$", "%"):
        text = text.replace(token, "")
    return Decimal(text.strip())


def _to_bool(raw: str) -> bool | None:
    """Parse a yes/no cell; None means the value was not recognised."""
    token = raw.strip().lower()
    if token in TRUE_TOKENS:
        return True
    if token in FALSE_TOKENS:
        return False
    return None


# ---------------------------------------------------------------------------
# Reading the file
# ---------------------------------------------------------------------------


def _normalise_header(text: str) -> str:
    """
    Reduce a header cell to a comparison key.

    Accents, case and punctuation vary with whoever retyped the header, and
    none of them carry meaning — "Prix de vente TTC" and "prix de vente ttc"
    are the same column.
    """
    import unicodedata

    decomposed = unicodedata.normalize("NFKD", text or "")
    stripped = "".join(char for char in decomposed if not unicodedata.combining(char))
    return "".join(char for char in stripped.lower() if char.isalnum())


def _header_lookup() -> dict[str, str]:
    """Map every accepted header spelling onto its column key."""
    lookup: dict[str, str] = {}
    for column in COLUMNS:
        # The key itself is accepted so a file produced by a developer, or a
        # re-import of an API export, works without renaming anything.
        for spelling in (column.header_fr, column.header_en, column.key):
            lookup[_normalise_header(spelling)] = column.key
    return lookup


def _read_rows(upload) -> tuple[list[str], list[list[str]]]:
    """
    Extract the header row and data rows from an uploaded file.

    Returns raw strings; interpretation belongs to `parse_workbook`.
    """
    name = (getattr(upload, "name", "") or "").lower()

    if name.endswith(".csv") or name.endswith(".txt"):
        raw = upload.read()
        if isinstance(raw, bytes):
            # utf-8-sig strips the BOM Excel writes; latin-1 is the fallback
            # for a file saved as "CSV (MS-DOS)" on a French Windows.
            try:
                text = raw.decode("utf-8-sig")
            except UnicodeDecodeError:
                text = raw.decode("latin-1")
        else:
            text = raw
        # Sniffing rather than assuming: this deployment writes semicolons,
        # but a file from an anglophone system uses commas and the user has no
        # way to tell which they have.
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ";" if sample.count(";") >= sample.count(",") else ","
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        table = [[_clean(cell) for cell in row] for row in reader]
        if not table:
            return [], []
        return table[0], table[1:]

    from openpyxl import load_workbook

    # read_only keeps a large file off the heap; data_only reads the cached
    # result of a formula rather than the formula text, which is what a user
    # who computed their prices in-sheet actually means.
    workbook = load_workbook(upload, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header = [_clean(cell) for cell in next(rows)]
        except StopIteration:
            return [], []
        data = [[_clean(cell) for cell in row] for row in rows]
    finally:
        # read_only mode holds the archive open until closed explicitly.
        workbook.close()

    return header, data


# ---------------------------------------------------------------------------
# Parsing and validation
# ---------------------------------------------------------------------------


def _resolve_reference(model, code: str, cache: dict):
    """
    Look up a reference record by code, memoised across the file.

    A catalogue import repeats the same handful of categories on every row;
    without the cache this is one query per cell.
    """
    if code in cache:
        return cache[code]
    # Case-insensitive because a code is a human-typed identifier here, not a
    # generated one — "ANALG" and "analg" are the same category to the user.
    instance = model.objects.filter(
        code__iexact=code, deleted_at__isnull=True
    ).first()
    cache[code] = instance
    return instance


def parse_workbook(upload, *, language: str = "fr") -> ImportReport:
    """
    Read and validate an uploaded catalogue file without writing anything.

    Every row is checked independently, so one bad row does not hide the state
    of the rest — the operator gets a single complete list of what to fix.
    """
    report = ImportReport()

    try:
        header, data_rows = _read_rows(upload)
    except Exception as exc:  # noqa: BLE001 - surfaced to the user verbatim
        report.fatal_error = _(
            "The file could not be read. Check that it is a valid .xlsx or .csv file. (%(detail)s)"
        ) % {"detail": str(exc)[:200]}
        return report

    if not header:
        report.fatal_error = _("The file is empty.")
        return report

    lookup = _header_lookup()
    # Position of each recognised column. Unknown columns are ignored rather
    # than rejected: an exported legacy file carries columns this catalogue has
    # no field for, and refusing the whole file over them helps nobody.
    positions: dict[str, int] = {}
    for index, cell in enumerate(header):
        key = lookup.get(_normalise_header(cell))
        if key is not None and key not in positions:
            positions[key] = index

    missing = [
        column.header(language)
        for column in COLUMNS
        if column.required and column.key not in positions
    ]
    if missing:
        report.fatal_error = _(
            "These required columns are missing from the file: %(columns)s. "
            "Download the template to get the expected layout."
        ) % {"columns": ", ".join(missing)}
        return report

    if len(data_rows) > MAX_ROWS:
        report.fatal_error = _(
            "The file contains %(count)s rows, more than the %(max)s allowed in one import. "
            "Split it into several files."
        ) % {"count": len(data_rows), "max": MAX_ROWS}
        return report

    category_cache: dict[str, Category | None] = {}
    manufacturer_cache: dict[str, Manufacturer | None] = {}
    unit_cache: dict[str, UnitOfMeasure | None] = {}

    valid_forms = set(DosageForm.values)
    valid_storage = set(StorageCondition.values)

    # Duplicate detection is per-file as well as against the database: two rows
    # claiming the same product code would otherwise both pass validation and
    # the second would fail at write time, after the first was committed.
    seen_codes: dict[str, int] = {}
    seen_barcodes: dict[str, int] = {}

    for offset, raw_row in enumerate(data_rows):
        # +2: one for the header row, one because Excel numbers from 1.
        row_number = offset + 2

        def cell(key: str) -> str:
            index = positions.get(key)
            if index is None or index >= len(raw_row):
                return ""
            return raw_row[index]

        # A trailing run of blank rows is normal in a hand-edited sheet and is
        # silently skipped; a row with any content at all is validated.
        if not any(cell(column.key) for column in COLUMNS):
            continue

        result = RowResult(row_number=row_number)
        errors = result.errors
        data = result.data

        def add_error(key: str, message: str) -> None:
            errors.setdefault(key, []).append(message)

        # --- Text columns -------------------------------------------------
        for key in (
            "product_code", "name", "generic_name", "brand_name", "strength",
            "pack_size", "atc_code", "registration_number", "barcode", "notes",
        ):
            value = cell(key)
            if value:
                data[key] = value

        if not data.get("name"):
            add_error("name", _("This field is required."))

        # --- Choice columns -----------------------------------------------
        dosage_form = cell("dosage_form").upper().strip()
        if dosage_form:
            if dosage_form not in valid_forms:
                add_error(
                    "dosage_form",
                    _("'%(value)s' is not a valid dosage form. Accepted codes: %(codes)s.")
                    % {"value": dosage_form, "codes": ", ".join(sorted(valid_forms))},
                )
            else:
                data["dosage_form"] = dosage_form

        storage = cell("storage_condition").upper().strip()
        if storage:
            if storage not in valid_storage:
                add_error(
                    "storage_condition",
                    _("'%(value)s' is not a valid storage condition. Accepted codes: %(codes)s.")
                    % {"value": storage, "codes": ", ".join(sorted(valid_storage))},
                )
            else:
                data["storage_condition"] = storage

        # --- Reference columns --------------------------------------------
        category_code = cell("category")
        if not category_code:
            add_error("category", _("This field is required."))
        else:
            category = _resolve_reference(Category, category_code, category_cache)
            if category is None:
                add_error(
                    "category",
                    _("No category has the code '%(code)s'. Create it first, or correct the code.")
                    % {"code": category_code},
                )
            else:
                data["category"] = category

        unit_code = cell("unit_of_measure")
        if not unit_code:
            add_error("unit_of_measure", _("This field is required."))
        else:
            unit = _resolve_reference(UnitOfMeasure, unit_code, unit_cache)
            if unit is None:
                add_error(
                    "unit_of_measure",
                    _("No unit has the code '%(code)s'. Create it first, or correct the code.")
                    % {"code": unit_code},
                )
            else:
                data["unit_of_measure"] = unit

        manufacturer_code = cell("manufacturer")
        if manufacturer_code:
            manufacturer = _resolve_reference(
                Manufacturer, manufacturer_code, manufacturer_cache
            )
            if manufacturer is None:
                add_error(
                    "manufacturer",
                    _("No manufacturer has the code '%(code)s'.")
                    % {"code": manufacturer_code},
                )
            else:
                data["manufacturer"] = manufacturer

        # --- Numeric columns ----------------------------------------------
        for key, default in DECIMAL_DEFAULTS.items():
            raw = cell(key)
            if not raw:
                data[key] = default
                continue
            try:
                value = _to_decimal(raw)
            except (InvalidOperation, ValueError, ArithmeticError):
                add_error(
                    key,
                    _("'%(value)s' is not a number.") % {"value": raw},
                )
                continue
            if value < 0:
                add_error(key, _("This value cannot be negative."))
                continue
            data[key] = value

        raw_days = cell("expiry_alert_days")
        if raw_days:
            try:
                days = int(_to_decimal(raw_days))
            except (InvalidOperation, ValueError, ArithmeticError):
                add_error(
                    "expiry_alert_days",
                    _("'%(value)s' is not a whole number of days.") % {"value": raw_days},
                )
            else:
                if days < 0:
                    add_error("expiry_alert_days", _("This value cannot be negative."))
                else:
                    data["expiry_alert_days"] = days

        # Mirrors the model's check constraint, which would otherwise surface
        # as an opaque IntegrityError halfway through the write.
        maximum = data.get("max_stock_level")
        reorder = data.get("reorder_level")
        if (
            maximum is not None
            and reorder is not None
            and maximum > 0
            and maximum < reorder
        ):
            add_error(
                "max_stock_level",
                _("The maximum stock level must be above the reorder level."),
            )

        # --- Boolean columns ----------------------------------------------
        for key in BOOLEAN_COLUMNS:
            raw = cell(key)
            if not raw:
                data[key] = False
                continue
            parsed = _to_bool(raw)
            if parsed is None:
                add_error(
                    key,
                    _("'%(value)s' is not recognised. Use YES or NO.") % {"value": raw},
                )
            else:
                data[key] = parsed

        # --- Uniqueness ----------------------------------------------------
        # A product code that already exists is not an error: re-importing a
        # corrected file is the normal way to fix a catalogue, and the row is
        # applied as an update. What is rejected is the same code appearing
        # twice within one file, which is always a mistake.
        product_code = data.get("product_code", "")
        if product_code:
            if product_code in seen_codes:
                add_error(
                    "product_code",
                    _("This product code is already used on row %(row)s of this file.")
                    % {"row": seen_codes[product_code]},
                )
            else:
                seen_codes[product_code] = row_number
                existing = Medicine.objects.filter(
                    product_code=product_code, deleted_at__isnull=True
                ).values_list("id", flat=True).first()
                if existing is not None:
                    result.existing_id = str(existing)

        barcode = data.get("barcode", "")
        if barcode:
            if barcode in seen_barcodes:
                add_error(
                    "barcode",
                    _("This barcode is already used on row %(row)s of this file.")
                    % {"row": seen_barcodes[barcode]},
                )
            else:
                seen_barcodes[barcode] = row_number
                clash = (
                    Medicine.objects.filter(barcode=barcode, deleted_at__isnull=True)
                    .exclude(id=result.existing_id or None)
                    .exists()
                    if result.existing_id
                    else Medicine.objects.filter(
                        barcode=barcode, deleted_at__isnull=True
                    ).exists()
                )
                if clash:
                    add_error(
                        "barcode",
                        _("Another product already uses this barcode."),
                    )

        report.rows.append(result)

    if not report.rows:
        report.fatal_error = _("The file contains no data rows.")

    return report


# ---------------------------------------------------------------------------
# Writing
# ---------------------------------------------------------------------------


@transaction.atomic
def commit(report: ImportReport, *, actor=None) -> ImportReport:
    """
    Persist the valid rows of a parsed report.

    Atomic by design: an import that half-applied would leave the operator
    reconciling their spreadsheet against the catalogue row by row. Either the
    whole file lands or none of it does, and error rows are reported back so
    the file can be corrected and re-sent.

    Writes go through `services.create_medicine` / `update_medicine`, so an
    imported product is indistinguishable from one typed into the form: same
    product-code allocation, same opening price-history row, same audit entry.
    """
    from . import services

    for row in report.valid_rows:
        payload = dict(row.data)
        # `status` is not an import column: a spreadsheet describes what a
        # product is, not where it sits in its lifecycle. Everything imported
        # is an active catalogue line until someone changes it on the screen.
        payload.setdefault("status", ProductStatus.ACTIVE)

        if row.existing_id:
            medicine = Medicine.objects.get(pk=row.existing_id)
            # The code identifies the row; resending it as a field would make
            # the update look like a rename of itself.
            payload.pop("product_code", None)
            # An update that changes pricing needs a reason for the history
            # row, and "imported" is the honest one.
            payload["price_change_reason"] = str(_("Catalogue import"))
            services.update_medicine(medicine, actor=actor, **payload)
            report.updated += 1
        else:
            services.create_medicine(actor=actor, **payload)
            report.created += 1

    return report


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------


def build_template(*, language: str = "fr") -> bytes:
    """
    Generate the .xlsx import template as raw bytes.

    The template is the specification. It carries three things a written
    description cannot: the exact header text the parser matches, dropdown
    validation on the coded columns so an invalid dosage form cannot be typed
    in the first place, and a second sheet listing the category, unit and
    manufacturer codes that exist *in this database right now* — which is the
    part a static template can never be right about.
    """
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    english = language.startswith("en")
    workbook = Workbook()

    sheet = workbook.active
    sheet.title = "Import" if english else "Import"

    header_fill = PatternFill("solid", fgColor="0F5132")
    required_fill = PatternFill("solid", fgColor="B45309")
    header_font = Font(bold=True, color="FFFFFF")

    for index, column in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=index, value=column.header(language))
        # Required columns are amber, optional ones green: the operator can see
        # the minimum viable file at a glance without reading any instructions.
        cell.fill = required_fill if column.required else header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        hint = column.hint(language)
        if column.required:
            hint = (
                f"{'Required. ' if english else 'Obligatoire. '}{hint}".strip()
            )
        if hint:
            # A cell comment rather than a separate instructions row: it is
            # visible on hover and, crucially, does not shift the data down a
            # row and break the parser's expectation of headers on row 1.
            cell.comment = Comment(hint, "Vision Pharma plus")

        sheet.column_dimensions[get_column_letter(index)].width = column.width

    # One worked example so the expected shape of each cell is unambiguous.
    # It is styled as a hint and labelled, so it is obvious it must be replaced
    # rather than left in.
    example_font = Font(italic=True, color="6B7280")
    for index, column in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=2, column=index, value=column.example or None)
        cell.font = example_font

    sheet.freeze_panes = "A3"

    # --- Reference sheet --------------------------------------------------
    reference = workbook.create_sheet("Références" if not english else "Reference")
    reference_row = 1

    def write_block(title: str, pairs: list[tuple[str, str]]) -> tuple[int, int]:
        """Write a titled two-column block; returns its first and last data row."""
        nonlocal reference_row
        cell = reference.cell(row=reference_row, column=1, value=title)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        reference.cell(row=reference_row, column=2).fill = header_fill
        reference_row += 1
        first = reference_row
        for code, label in pairs:
            reference.cell(row=reference_row, column=1, value=code)
            reference.cell(row=reference_row, column=2, value=label)
            reference_row += 1
        last = reference_row - 1
        reference_row += 1  # blank spacer between blocks
        return first, last

    categories = list(
        Category.objects.filter(is_active=True, deleted_at__isnull=True)
        .order_by("code")
        .values_list("code", "name_en" if english else "name_fr")
    )
    units = list(
        UnitOfMeasure.objects.filter(is_active=True, deleted_at__isnull=True)
        .order_by("code")
        .values_list("code", "name_en" if english else "name_fr")
    )
    manufacturers = list(
        Manufacturer.objects.filter(is_active=True, deleted_at__isnull=True)
        .order_by("code")
        .values_list("code", "name")
    )

    category_range = write_block(
        "Category codes" if english else "Codes catégorie", categories
    )
    unit_range = write_block("Unit codes" if english else "Codes unité", units)
    write_block(
        "Manufacturer codes" if english else "Codes fabricant", manufacturers
    )
    write_block(
        "Dosage forms" if english else "Formes galéniques",
        [(value, str(label)) for value, label in DosageForm.choices],
    )
    write_block(
        "Storage conditions" if english else "Conditions de conservation",
        [(value, str(label)) for value, label in StorageCondition.choices],
    )

    reference.column_dimensions["A"].width = 22
    reference.column_dimensions["B"].width = 46

    # --- Dropdowns --------------------------------------------------------
    # Applied over a generous row range so a user pasting 800 rows still gets
    # validation. Excel evaluates these lazily, so the range costs nothing.
    last_row = 1000

    def add_dropdown(column_key: str, formula: str) -> None:
        index = next(
            (i for i, column in enumerate(COLUMNS, start=1) if column.key == column_key),
            None,
        )
        if index is None:
            return
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        # showErrorMessage on but not "stop": a code that is legitimate and
        # simply newer than the template must still be typeable. The dropdown
        # is an aid, and the server remains the authority.
        validation.errorStyle = "warning"
        validation.error = (
            "This code is not in the reference list. Check the 'Reference' tab."
            if english
            else "Ce code ne figure pas dans la liste. Voir l'onglet « Références »."
        )
        sheet.add_data_validation(validation)
        letter = get_column_letter(index)
        validation.add(f"{letter}3:{letter}{last_row}")

    sheet_name = f"'{reference.title}'"
    if categories:
        add_dropdown(
            "category",
            f"={sheet_name}!$A${category_range[0]}:$A${category_range[1]}",
        )
    if units:
        add_dropdown(
            "unit_of_measure", f"={sheet_name}!$A${unit_range[0]}:$A${unit_range[1]}"
        )
    add_dropdown("dosage_form", f'"{",".join(DosageForm.values)}"')
    add_dropdown("storage_condition", f'"{",".join(StorageCondition.values)}"')

    yes_no = "YES,NO" if english else "OUI,NON"
    for key in BOOLEAN_COLUMNS:
        add_dropdown(key, f'"{yes_no}"')

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
